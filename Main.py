import math
import openpyxl
import random
from random import randint
from datetime import datetime, timedelta
import os

baslangic_tarihi = datetime(2023, randint(1, 3), randint(1, 28), randint(1, 22))


def olustur_ve_aktif_et(dosya_adi):
    # Dosya var mı kontrol et
    if not os.path.exists(dosya_adi):
        # Dosya yoksa, yeni bir Excel dosyası oluştur
        workbook = openpyxl.Workbook()
        workbook.save(dosya_adi)
        return dosya_adi
    else:
        # Dosya varsa, dosya adını incele
        dosya_sayisi = 1
        while f"Data{dosya_sayisi}.xlsx" in os.listdir():
            dosya_sayisi += 1

        # Yeni dosya adını oluştur
        yeni_dosya_adi = f"Data{dosya_sayisi}.xlsx"

        # Yeni dosyayı oluştur
        workbook = openpyxl.Workbook()
        workbook.save(yeni_dosya_adi)

        return yeni_dosya_adi


def addData(pageName, firstStep, secondStep, startSpeed, endSpeed, sensitivity):
    for row in range(firstStep, secondStep):
        randomValue = random.uniform(startSpeed, endSpeed) + row / sensitivity
        speedChange = randomValue * 2 * math.pi
        energy = 0.5 * (0.5 * 180 * 0.2 * 0.2) * speedChange * speedChange
        pageName.cell(row=row, column=1, value=str(row - 1) + ". Veri")
        pageName.cell(row=row, column=2, value=randomValue)
        pageName.cell(row=row, column=3, value=speedChange)
        pageName.cell(row=row, column=4, value=energy)
        pageName.cell(row=row, column=5, value=getTime(randomValue))


def getTime(time):
    global baslangic_tarihi

    # 250'de 1 ihtimal gerçekleşirse 1,2 veya 3 saat molaver
    number = randint(1, 250)
    if (number == 150):
        baslangic_tarihi += timedelta(hours=randint(1, 4))
        if baslangic_tarihi.hour > 17:
            baslangic_tarihi += timedelta(hours=15)

    # Cumartesi günü ise günü 2 artır
    if baslangic_tarihi.weekday() == 5:
        baslangic_tarihi += timedelta(days=2)
    # Pazar günü ise günü 1 artır
    elif baslangic_tarihi.weekday() == 6:
        baslangic_tarihi += timedelta(days=1)

    # Saat 17:00 ise saati 15 saat ekleyerek 8:00'e getir
    if baslangic_tarihi.hour == 17:
        baslangic_tarihi += timedelta(hours=15)

    milis = time * 1000
    baslangic_tarihi += timedelta(milliseconds=milis)

    return baslangic_tarihi

def get_latest_data_file():
    data_files = [file for file in os.listdir() if file.startswith("Data") and file.endswith(".xlsx")]

    if not data_files:
        return None  # Eğer uygun dosya bulunamazsa None döndür

    latest_file = max(data_files, key=os.path.getctime)
    return latest_file

def generateAndSaveData(entire=True,isEstimated=False,dataList=[]):

    if isEstimated:

        fileName=get_latest_data_file()
        workbook = openpyxl.load_workbook(fileName)
        yeni_sayfa = workbook.active
        addEstimatedData(yeni_sayfa, dataList)
        workbook.save(fileName)
        return



    fileName = "Data1.xlsx"

    fileName = olustur_ve_aktif_et(fileName)

    workbook = openpyxl.load_workbook(fileName)

    yeni_sayfa = workbook.create_sheet("Data")
    workbook.active = yeni_sayfa

    # Başlık ekle
    yeni_sayfa['A1'] = "VERİ"
    yeni_sayfa['B1'] = 'DEVİR DEĞİŞİMİ'
    yeni_sayfa['C1'] = 'HIZ DEĞİŞİMİ'
    yeni_sayfa['D1'] = 'ENERJİ'
    yeni_sayfa['E1'] = 'ZAMAN DAMGASI'

    firstStep = randint(70000, 85000)
    addData(yeni_sayfa, 2, firstStep, 1.1, 1.4, 800000)
    if entire==False:
        workbook.save(fileName)
        return


    secondStep = firstStep + randint(10000, 18000)
    addData(yeni_sayfa, firstStep, secondStep, 1.20, 1.49, 900000)

    thirdStep = secondStep + randint(6000, 10000)
    addData(yeni_sayfa, secondStep, thirdStep, 1.34, 1.56, 1000000)

    forthStep = thirdStep + randint(50, 250)
    addData(yeni_sayfa, thirdStep, forthStep, 1.45, 1.95, 500000)

    # Excel dosyasını kaydet
    workbook.save(fileName)


def addEstimatedData(pageName, dataList):
    for row in range(2, len(dataList)):
        pageName.cell(row=row, column=2, value=dataList[row-2])


def updateLastFileWithEstimatedData(dataList):
    generateAndSaveData(True, True, dataList)


def addNewData(times,entireOrNot):
    for i in range(times):
        generateAndSaveData(entireOrNot)
        baslangic_tarihi = datetime(2023, randint(1, 3), randint(1, 28), randint(1, 22))